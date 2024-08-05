from openpyxl.reader.excel import load_workbook
from selenium.webdriver import ActionChains


class TestUtil:

    @staticmethod
    def mouse_hover(driver, web_element):
        action = ActionChains(driver)
        action.move_to_element(web_element).perform()

    @staticmethod
    def read_excel_data(filepath, sheet_name):
        workbook = load_workbook(filename=filepath)
        sheet = workbook[sheet_name]
        test_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            test_data.append(row)
        return test_data